#Based on https://qiita.com/matoarea/items/af802ffe430d5ce364fd

import os
import openpyxl
from openpyxl import Workbook
import time
from pymeasure.instruments.keithley import Keithley2400
from datetime import datetime

date = int(datetime.now().strftime("%Y%m%d"))-20000000 # 日付。2020年9月10日なら「200910」
keithley = Keithley2400("GPIB::24") # 使用する機器に該当するGPIBアドレスを指定。
book = openpyxl.Workbook() # エクセルファイルを作成
sheet = book.worksheets[0] # エクセルファイル中の使用するブックを指定。

x = [] # voltage
y = [] # current
z = [] # time

def initial_settings():

    # エクセルファイルの作成
    sheet.cell(row = 1, column = 1).value = 'voltage (V)'
    sheet.cell(row = 1, column = 2).value = 'current (mA)'
    sheet.cell(row = 1, column = 3).value = 'time (sec)'
    sheet.cell(row = 1, column = 4).value = datetime.now()

    # ソースメータの設定
    keithley.reset() # とりあえずreset
    keithley.disable_buffer()
    keithley.use_front_terminals()
    keithley.apply_voltage() # 電圧印可モード。
    keithley.source_voltage_range = 1 # 印可する電圧のレンジ。 1.000 V のレンジで印可。
    keithley.source_voltage = 0 # ソース電圧（印可される電圧）の大きさをとりあえず0に。
    keithley.enable_source() # 印可を行う（ソース電源の印可を有効にする）。
    keithley.measure_current(nplc=0.01, current=0.000105, auto_range=True) # 電流測定。測定に用いる積分時間の長さを0.01 nplcに設定。測定される電流のレンジを
    keithley.current_range = I_sensitivity
    keithley.compliance_current = 0.01 # まず超えないが、一応設定。
    keithley.wires = 4 # 4線式で行うことが肝要。理由も説明できるようになるべし。

def voltage_apply(voltage):
    measurement_interval = 1/(scan_rate*1000) #余談だが、0.001/scan_rateにすべきではない。小数を用いると桁落ちの危険。
    meas_time = measurement_interval
    time_accuracy = 0.001

    points = int((high_V-low_V)*1000) # the number of point which should be recorded. I value is recorded at every 0.001 V.
    dV = 0.001 # 掃引電圧の粗さ 兼 （今回の場合は）電流測定間隔
    for i in range(segments):
        step = 0
        for step in range (points): 
            keithley.source_voltage = voltage
            rest_measurement = time.time() - base_time - meas_time
            while rest_measurement < 0: # 電流反転予定の時間が来るまで待つ。
                rest_measurement = time.time() - base_time - meas_time
                time.sleep(time_accuracy)
            z.append(time.time()-base_time) #時間記録
            x.append(voltage) # 電圧記録
            y.append(keithley.current) # 電流記録(A)
            voltage = voltage + dV # 印可電圧の変更（反映されるのは次のループから）
            meas_time = meas_time + measurement_interval # 次の測定点を設定
        dV = dV*(-1) # 掃引方向を変更
    keithley.shutdown()

    # エクセルにデータ記入。リストに一旦入れた値をエクセルへ。
    #この方法が毎度毎度エクセルに記録するよりも速いのかは未検証。
    for i in range(segments * points): 
        sheet.cell(row = 2+i, column = 1).value = x[i]
        sheet.cell(row = 2+i, column = 2).value = y[i]*1000 # 電流をmA単位で記入する。
        sheet.cell(row = 2+i, column = 3).value = z[i]
    print('voltage_apply_done.')

def duplicate_rename(file_path):
    if os.path.exists(file_path):
        name, ext = os.path.splitext(file_path)
        i = 1
        while True:
            # 数値を3桁などにしたい場合は({:0=3})とする
            new_name = "{} ({:0=2}){}".format(name, i, ext)
            if not os.path.exists(new_name):
                return new_name
            i += 1
    else:
        return file_path
    
### conducting functions ###
if __name__ == "__main__": # おまじないみたいなもの。気になるならググって。
    # ファイル情報
    path = "C:\\Users\\（各自のユーザー名）\\Desktop\\" #ファイル保存場所を好きに設定。
    #とりあえずデスクトップに保存することにしてみた。
    sample_name = 'test' # ここに入力したサンプル名がファイル名に反映される。

    # 測定条件
    high_V = 0.5 # high voltage (V)
    low_V = 0.4 # low voltage (V)
    start_V = low_V # 必ずしも start_V = low_V ではないが、今回は簡単のため一致させている。
    finish_V = high_V
    #direction = n # p/nの設定で掃引方向を設定できるようにしましょう。
    scan_rate = 0.05 # scan rate(V/s)
    segments = 2 # number of segments (2 segments is equal to 1 cycle )
    I_sensitivity = 0.0001 # sensitivity of current measurement (A). 0.0001 means 0.00105 uA is maximum.

    # 保存ファイル名を生成
    filename = "{0}_{1}_{2}-{3}V_{4}Vs-1.xlsx".format(date, sample_name, low_V,high_V, scan_rate) 
    filename = duplicate_rename(filename)

    # 測定の実行
    initial_settings()
    keithley.enable_source()# なぜかここでもoutputをenableしないとoutputがenableにならない
    base_time = time.time() # 測定において0秒となる時刻を取得
    voltage_apply(start_V)
    book.save( path + filename ) # エクセルファイルを filename の名前で、path の場所に保存。
    print('finished.') 
